from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableView, 
    QGroupBox, QFormLayout, QDateEdit, QComboBox, QLabel, QFileDialog
)
from PyQt6.QtCore import QDate
from PyQt6.QtGui import QStandardItemModel, QStandardItem
import openpyxl
from openpyxl.styles import Font, Alignment
from fpdf import FPDF
import os
from datetime import datetime
from database import init_db
from sqlalchemy.orm import sessionmaker
from sqlalchemy import func, extract
from database.models import Payment, Order, Client

class ReportForm(QWidget):
    def __init__(self):
        super().__init__()
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)
        
        # Фильтры
        self.filter_group = QGroupBox("Параметры отчета")
        self.filter_layout = QFormLayout()
        self.filter_group.setLayout(self.filter_layout)
        
        self.report_type = QComboBox()
        self.report_type.addItems([
            "Прибыль по месяцам", 
            "Прибыль по клиентам", 
            "Активность перевозчиков"
        ])
        
        self.year_combo = QComboBox()
        current_year = QDate.currentDate().year()
        for year in range(current_year - 5, current_year + 1):
            self.year_combo.addItem(str(year), year)
        self.year_combo.setCurrentText(str(current_year))
        
        self.client_combo = QComboBox()
        self.generate_button = QPushButton("Сформировать отчет")
        self.generate_button.clicked.connect(self.generate_report)
        
        # Кнопки экспорта
        self.export_layout = QHBoxLayout()
        self.export_excel_button = QPushButton("Экспорт в Excel")
        self.export_excel_button.clicked.connect(self.export_to_excel)
        self.export_pdf_button = QPushButton("Экспорт в PDF")
        self.export_pdf_button.clicked.connect(self.export_to_pdf)
        
        self.export_layout.addWidget(self.export_excel_button)
        self.export_layout.addWidget(self.export_pdf_button)
        
        self.filter_layout.addRow("Тип отчета:", self.report_type)
        self.filter_layout.addRow("Год:", self.year_combo)
        self.filter_layout.addRow("Клиент (если применимо):", self.client_combo)
        self.filter_layout.addRow(self.generate_button)
        
        # Таблица отчета
        self.table = QTableView()
        self.model = QStandardItemModel()
        self.table.setModel(self.model)
        
        # Итоги
        self.summary_label = QLabel()
        self.summary_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        
        # Компоновка
        self.layout.addWidget(self.filter_group)
        self.layout.addWidget(self.table)
        self.layout.addWidget(self.summary_label)
        self.layout.addLayout(self.export_layout)
        
        # Инициализация БД
        self.engine = init_db()
        self.Session = sessionmaker(bind=self.engine)
        self.load_clients()
    
    def load_clients(self):
        session = self.Session()
        clients = session.query(Client).order_by(Client.name).all()
        self.client_combo.addItem("Все клиенты", 0)
        for client in clients:
            self.client_combo.addItem(client.name, client.id)
    
    def generate_report(self):
        report_type = self.report_type.currentText()
        year = self.year_combo.currentData()
        
        if report_type == "Прибыль по месяцам":
            self.generate_monthly_profit_report(year)
        elif report_type == "Прибыль по клиентам":
            self.generate_client_profit_report(year)
        elif report_type == "Активность перевозчиков":
            self.generate_carrier_activity_report(year)
    
    def generate_monthly_profit_report(self, year):
        session = self.Session()
        
        # Запрос для расчета прибыли по месяцам
        results = session.query(
            extract('month', Payment.payment_date).label('month'),
            func.sum(func.case(
                [(Payment.is_client_payment == True, Payment.amount)],
                else_=0
            )).label('income'),
            func.sum(func.case(
                [(Payment.is_client_payment == False, Payment.amount)],
                else_=0
            )).label('expense'),
            (func.sum(func.case(
                [(Payment.is_client_payment == True, Payment.amount)],
                else_=0
            )) - func.sum(func.case(
                [(Payment.is_client_payment == False, Payment.amount)],
                else_=0
            ))).label('profit')
        ).filter(
            extract('year', Payment.payment_date) == year
        ).group_by(
            extract('month', Payment.payment_date)
        ).order_by(
            extract('month', Payment.payment_date)
        ).all()
        
        # Настройка модели таблицы
        self.model.clear()
        self.model.setHorizontalHeaderLabels([
            "Месяц", "Доход", "Расход", "Прибыль"
        ])
        
        total_income = 0
        total_expense = 0
        total_profit = 0
        
        for row, (month, income, expense, profit) in enumerate(results):
            self.model.insertRow(row)
            self.model.setItem(row, 0, QStandardItem(self.get_month_name(month)))
            self.model.setItem(row, 1, QStandardItem(f"{income:.2f} руб."))
            self.model.setItem(row, 2, QStandardItem(f"{expense:.2f} руб."))
            self.model.setItem(row, 3, QStandardItem(f"{profit:.2f} руб."))
            
            total_income += income
            total_expense += expense
            total_profit += profit
        
        self.summary_label.setText(
            f"Итого за {year} год: "
            f"Доход: {total_income:.2f} руб., "
            f"Расход: {total_expense:.2f} руб., "
            f"Прибыль: {total_profit:.2f} руб."
        )
    
    def get_month_name(self, month_num):
        months = [
            "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
            "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
        ]
        return months[int(month_num) - 1] if 1 <= month_num <= 12 else str(month_num)
    
    def generate_client_profit_report(self, year):
        session = self.Session()
        
        # Запрос для расчета прибыли по клиентам
        results = session.query(
            Client.name,
            func.sum(func.case(
                [(Payment.is_client_payment == True, Payment.amount)],
                else_=0
            )).label('income'),
            func.sum(func.case(
                [(Payment.is_client_payment == False, Payment.amount)],
                else_=0
            )).label('expense'),
            (func.sum(func.case(
                [(Payment.is_client_payment == True, Payment.amount)],
                else_=0
            )) - func.sum(func.case(
                [(Payment.is_client_payment == False, Payment.amount)],
                else_=0
            ))).label('profit')
        ).join(Order).join(Payment).filter(
            extract('year', Payment.payment_date) == year
        ).group_by(
            Client.name
        ).order_by(
            func.sum(func.case(
                [(Payment.is_client_payment == True, Payment.amount)],
                else_=0
            )) - func.sum(func.case(
                [(Payment.is_client_payment == False, Payment.amount)],
                else_=0
            )).desc()
        ).all()
        
        # Настройка модели таблицы
        self.model.clear()
        self.model.setHorizontalHeaderLabels([
            "Клиент", "Доход", "Расход", "Прибыль"
        ])
        
        total_income = 0
        total_expense = 0
        total_profit = 0
        
        for row, (client_name, income, expense, profit) in enumerate(results):
            self.model.insertRow(row)
            self.model.setItem(row, 0, QStandardItem(client_name))
            self.model.setItem(row, 1, QStandardItem(f"{income:.2f} руб."))
            self.model.setItem(row, 2, QStandardItem(f"{expense:.2f} руб."))
            self.model.setItem(row, 3, QStandardItem(f"{profit:.2f} руб."))
            
            total_income += income
            total_expense += expense
            total_profit += profit
        
        self.summary_label.setText(
            f"Итого за {year} год: "
            f"Доход: {total_income:.2f} руб., "
            f"Расход: {total_expense:.2f} руб., "
            f"Прибыль: {total_profit:.2f} руб."
        )
    
    def generate_carrier_activity_report(self, year):
        session = self.Session()
        
        # Запрос для расчета активности перевозчиков
        results = session.query(
            Carrier.company_name,
            func.count(Order.id).label('order_count'),
            func.sum(Payment.amount).filter(Payment.is_client_payment == False).label('total_payments')
        ).join(Order).join(Payment).filter(
            extract('year', Order.order_date) == year,
            Payment.is_client_payment == False
        ).group_by(
            Carrier.company_name
        ).order_by(
            func.count(Order.id).desc()
        ).all()
        
        # Настройка модели таблицы
        self.model.clear()
        self.model.setHorizontalHeaderLabels([
            "Перевозчик", "Количество заказов", "Выплачено"
        ])
        
        total_orders = 0
        total_payments = 0
        
        for row, (carrier_name, order_count, payments) in enumerate(results):
            self.model.insertRow(row)
            self.model.setItem(row, 0, QStandardItem(carrier_name))
            self.model.setItem(row, 1, QStandardItem(str(order_count)))
            self.model.setItem(row, 2, QStandardItem(f"{payments:.2f} руб." if payments else "0.00 руб."))
            
            total_orders += order_count
            total_payments += payments if payments else 0
        
        self.summary_label.setText(
            f"Итого за {year} год: "
            f"Заказов: {total_orders}, "
            f"Выплачено перевозчикам: {total_payments:.2f} руб."
        )
    
    def export_to_excel(self):
        if self.model.rowCount() == 0:
            QMessageBox.warning(self, "Ошибка", "Нет данных для экспорта")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить отчет", "", "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Отчет"
            
            # Заголовки
            headers = []
            for col in range(self.model.columnCount()):
                headers.append(self.model.headerData(col, 1))  # 1 - горизонтальная ориентация
            
            ws.append(headers)
            
            # Жирный шрифт для заголовков
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Данные
            for row in range(self.model.rowCount()):
                row_data = []
                for col in range(self.model.columnCount()):
                    item = self.model.item(row, col)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)
            
            # Автоширина столбцов
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(file_path)
            QMessageBox.information(self, "Успех", "Отчет успешно экспортирован в Excel")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл: {str(e)}")
    
    def export_to_pdf(self):
        if self.model.rowCount() == 0:
            QMessageBox.warning(self, "Ошибка", "Нет данных для экспорта")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить отчет", "", "PDF Files (*.pdf)"
        )
        
        if not file_path:
            return
        
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=10)
            
            # Заголовок отчета
            report_type = self.report_type.currentText()
            year = self.year_combo.currentText()
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 10, f"Отчет: {report_type} за {year} год", 0, 1, 'C')
            pdf.ln(5)
            
            # Заголовки таблицы
            pdf.set_font("Arial", 'B', 10)
            headers = []
            col_width = 190 / self.model.columnCount()  # Равномерное распределение
            
            for col in range(self.model.columnCount()):
                header = self.model.headerData(col, 1)  # 1 - горизонтальная ориентация
                headers.append(header)
                pdf.cell(col_width, 10, header, 1, 0, 'C')
            
            pdf.ln()
            
            # Данные таблицы
            pdf.set_font("Arial", size=10)
            for row in range(self.model.rowCount()):
                for col in range(self.model.columnCount()):
                    item = self.model.item(row, col)
                    text = item.text() if item else ""
                    pdf.cell(col_width, 10, text, 1, 0, 'C')
                pdf.ln()
            
            # Итоговая информация
            if self.summary_label.text():
                pdf.ln(5)
                pdf.set_font("Arial", 'B', 10)
                pdf.cell(0, 10, self.summary_label.text(), 0, 1)
            
            # Дата создания
            pdf.ln(10)
            pdf.set_font("Arial", 'I', 8)
            pdf.cell(0, 10, f"Создано: {datetime.now().strftime('%d.%m.%Y %H:%M')}", 0, 1, 'R')
            
            pdf.output(file_path)
            QMessageBox.information(self, "Успех", "Отчет успешно экспортирован в PDF")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл: {str(e)}")